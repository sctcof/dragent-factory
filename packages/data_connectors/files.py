from __future__ import annotations

import csv
import re
from collections import Counter
from pathlib import Path
from typing import Any, Dict, List

from packages.shared_types.models import DataDictionary, FieldProfile, KnowledgeGraph


EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
PHONE_RE = re.compile(r"^\+?\d[\d\s-]{7,}\d$")


def parse_uploaded_table(asset_id: str, asset_name: str, object_path: Path) -> Dict[str, Any]:
    rows = _read_rows(object_path)
    headers = list(rows[0].keys()) if rows else []
    profiles = [_profile_column(header, rows) for header in headers]
    metrics = [
        {
            "name": profile.name,
            "formula": f"sum({profile.name})",
            "derived_from": [profile.name],
            "supporting_ids": [f"{asset_id}:column:{profile.name}"],
        }
        for profile in profiles
        if profile.logical_type == "number"
    ]
    dictionary = DataDictionary(
        asset_id=asset_id,
        asset_name=asset_name,
        table_name=_table_name(asset_name),
        row_count=len(rows),
        columns=profiles,
        metrics=metrics,
        supporting_ids=[f"{asset_id}:schema", f"{asset_id}:profile"],
    )
    return {
        "rows": rows,
        "dictionary": dictionary,
        "graph": _build_graph(dictionary),
    }


def _read_rows(path: Path) -> List[Dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix in [".xlsx", ".xls"]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("Excel parsing requires openpyxl") from exc
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(value or "").strip() for value in rows[0]]
        result: List[Dict[str, Any]] = []
        for row in rows[1:]:
            result.append({headers[index]: value if value is not None else "" for index, value in enumerate(row) if index < len(headers) and headers[index]})
        return result

    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        return [dict(row) for row in reader]


def _profile_column(name: str, rows: List[Dict[str, Any]]) -> FieldProfile:
    values = [row.get(name, "") for row in rows]
    non_empty = [value for value in values if value not in ["", None]]
    converted_numbers = [_to_float(value) for value in non_empty]
    numeric_values = [value for value in converted_numbers if value is not None]
    logical_type = "number" if non_empty and len(numeric_values) / len(non_empty) >= 0.85 else "category"
    if logical_type != "number" and _looks_date(non_empty):
        logical_type = "date"
    sensitive = _looks_sensitive(name, non_empty)
    sample_values = []
    for value in non_empty:
        if value not in sample_values:
            sample_values.append(value)
        if len(sample_values) >= 5:
            break
    min_value = min(numeric_values) if numeric_values else None
    max_value = max(numeric_values) if numeric_values else None
    return FieldProfile(
        name=name,
        logical_type=logical_type,
        null_count=len(values) - len(non_empty),
        unique_count=len(set(str(value) for value in non_empty)),
        sample_values=sample_values,
        min_value=min_value,
        max_value=max_value,
        sensitive=sensitive,
    )


def _to_float(value: Any) -> float | None:
    try:
        return float(str(value).replace(",", ""))
    except (TypeError, ValueError):
        return None


def _looks_date(values: List[Any]) -> bool:
    if not values:
        return False
    matched = 0
    for value in values[:50]:
        text = str(value)
        if re.match(r"^\d{4}[-/]\d{1,2}([-/]\d{1,2})?", text):
            matched += 1
    return matched / min(len(values), 50) >= 0.6


def _looks_sensitive(name: str, values: List[Any]) -> bool:
    lower = name.lower()
    if any(token in lower for token in ["phone", "mobile", "email", "id_card", "身份证", "手机号", "邮箱"]):
        return True
    if not values:
        return False
    sample = [str(value) for value in values[:20]]
    return any(EMAIL_RE.match(value) or PHONE_RE.match(value) for value in sample)


def _table_name(asset_name: str) -> str:
    return Path(asset_name).stem.replace(" ", "_").lower()


def _build_graph(dictionary: DataDictionary) -> KnowledgeGraph:
    dataset_id = f"dataset:{dictionary.asset_id}"
    table_id = f"table:{dictionary.asset_id}:{dictionary.table_name}"
    nodes = [
        {"id": dataset_id, "type": "Dataset", "label": dictionary.asset_name},
        {"id": table_id, "type": "Table", "label": dictionary.table_name},
    ]
    edges = [{"source": dataset_id, "target": table_id, "type": "CONTAINS"}]
    for column in dictionary.columns:
        column_id = f"column:{dictionary.asset_id}:{column.name}"
        nodes.append({"id": column_id, "type": "Column", "label": column.name, "logical_type": column.logical_type})
        edges.append({"source": table_id, "target": column_id, "type": "CONTAINS"})
    for metric in dictionary.metrics:
        metric_id = f"metric:{dictionary.asset_id}:{metric['name']}"
        nodes.append({"id": metric_id, "type": "Metric", "label": metric["name"]})
        for source in metric["derived_from"]:
            edges.append({"source": metric_id, "target": f"column:{dictionary.asset_id}:{source}", "type": "DERIVED_FROM"})
    return KnowledgeGraph(nodes=nodes, edges=edges)


def summarize_rows(rows: List[Dict[str, Any]], max_rows: int = 25) -> List[Dict[str, Any]]:
    return rows[:max_rows]


def top_categories(rows: List[Dict[str, Any]], field: str, limit: int = 12) -> List[Dict[str, Any]]:
    counter = Counter(str(row.get(field, "")) for row in rows if row.get(field, "") not in ["", None])
    return [{"name": name, "count": count} for name, count in counter.most_common(limit)]
